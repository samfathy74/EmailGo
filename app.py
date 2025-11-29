from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file
from database import db, Contact, Template, Campaign, EmailLog, Reply, Settings, Server, ContactGroup, User
from sqlalchemy import text
from email_utils import send_email, check_replies
import os
import json
from openpyxl import load_workbook
import glob
import threading
from datetime import datetime, timedelta, timezone
from email.utils import parseaddr
import pytz
import shutil

basedir = os.path.abspath(os.path.dirname(__file__))
TEMPLATE_DIR = os.path.join(basedir, 'email_templates')

def get_file_templates():
    if not os.path.exists(TEMPLATE_DIR):
        os.makedirs(TEMPLATE_DIR)
    files = glob.glob(os.path.join(TEMPLATE_DIR, '*.html'))
    return [os.path.basename(f) for f in files]

from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from dotenv import load_dotenv

load_dotenv() # Load environment variables from .env file

app = Flask(__name__)
# Use environment variable for DB URL
app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get('DATABASE_URL')
if not app.config['SQLALCHEMY_DATABASE_URI']:
    # Fallback to local sqlite if no DATABASE_URL is set (optional, but good for safety)
    app.config['SQLALCHEMY_DATABASE_URI'] = r'sqlite:///email_marketing.db'

# Fix for some SQLAlchemy versions if the URL starts with postgres://
if app.config['SQLALCHEMY_DATABASE_URI'] and app.config['SQLALCHEMY_DATABASE_URI'].startswith("postgres://"):
    app.config['SQLALCHEMY_DATABASE_URI'] = app.config['SQLALCHEMY_DATABASE_URI'].replace("postgres://", "postgresql://", 1)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY')
db.init_app(app)

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

# UserMixin for the User model is now directly in database.py

@login_manager.user_loader
def load_user(user_id):
    return db.session.get(User, int(user_id))

with app.app_context():
    db.create_all()
    # Create default admin user if not exists
    if not User.query.filter_by(username='admin').first():
        admin = User(username='admin', password='adminpassword')
        db.session.add(admin)
        db.session.commit()
    
    # Migration for error_message column
    try:
        with db.engine.connect() as conn:
            conn.execute(text("ALTER TABLE campaign ADD COLUMN error_message TEXT"))
            conn.commit()
    except Exception:
        pass # Column likely exists

    # Reset stuck campaigns
    try:
        stuck_campaigns = Campaign.query.filter_by(status='sending').all()
        if stuck_campaigns:
            for campaign in stuck_campaigns:
                campaign.status = 'failed'
                campaign.error_message = "Campaign interrupted by server restart."
            db.session.commit()
            print(f"Reset {len(stuck_campaigns)} stuck campaigns to failed.")
    except Exception as e:
        print(f"Error resetting stuck campaigns: {e}")

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))

    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        next_url = request.form.get('next')
        
        user = User.query.filter_by(username=username).first()
        
        if user and user.password == password:
            login_user(user)
            flash('Login successful.', 'success')
            return redirect(next_url or url_for('dashboard'))
        else:
            flash('Invalid username or password.', 'error')
            
    return render_template('login.html', next=request.args.get('next'))

@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('You have been logged out.', 'success')
    return redirect(url_for('login'))

@app.route('/')
@login_required
def dashboard():
    # Fetch stats
    contact_count = Contact.query.filter_by(status='active').count()
    campaign_count = Campaign.query.count()
    
    # Advanced stats
    total_sent = EmailLog.query.filter_by(status='sent').count()
    total_failed = EmailLog.query.filter_by(status='failed').count()
    total_attempts = total_sent + total_failed
    success_rate = round((total_sent / total_attempts * 100), 1) if total_attempts > 0 else 0
    
    # Check replies logic removed to improve performance. 
    # Users can manually check for replies on the Replies page.
    reply_count = Reply.query.count()

    # Today's stats
    try:
        cairo_tz = pytz.timezone("Africa/Cairo")
        today = datetime.now(cairo_tz).date()
    except Exception as e:
        today = datetime.now(timezone.utc).date()
    sent_today = EmailLog.query.filter(db.func.date(EmailLog.sent_at) == today, EmailLog.status == 'sent').count()
    replies_today = Reply.query.filter(db.func.date(Reply.received_at) == today).count()

    # Recent activity
    recent_campaigns = Campaign.query.order_by(Campaign.created_at.desc()).limit(5).all()
    recent_replies = Reply.query.order_by(Reply.received_at.desc()).limit(5).all()

    # Chart Data: Last 7 Days Sent
    daily_stats_labels = []
    daily_stats_values = []
    for i in range(6, -1, -1):
        date = today - timedelta(days=i)
        count = EmailLog.query.filter(db.func.date(EmailLog.sent_at) == date, EmailLog.status == 'sent').count()
        daily_stats_labels.append(date.strftime('%d %b'))
        daily_stats_values.append(count)

    # Chart Data: Campaign Status
    campaign_statuses = db.session.query(Campaign.status, db.func.count(Campaign.status)).group_by(Campaign.status).all()
    campaign_stats = {status: count for status, count in campaign_statuses}
    # Ensure all keys exist for the chart
    status_order = ['completed', 'sending', 'failed', 'draft']
    campaign_stats_values = [campaign_stats.get(s, 0) for s in status_order]

    # Follow-ups sent
    followup_count = EmailLog.query.filter_by(status='sent', type='followup').count()

    # Failed and Draft Campaigns
    failed_campaign_count = Campaign.query.filter_by(status='failed').count()
    draft_campaign_count = Campaign.query.filter_by(status='draft').count()

    return render_template('dashboard.html', 
                           contact_count=contact_count, 
                           campaign_count=campaign_count, 
                           reply_count=reply_count,
                           followup_count=followup_count,
                           failed_campaign_count=failed_campaign_count,
                           draft_campaign_count=draft_campaign_count,
                           total_sent=total_sent,
                           success_rate=success_rate,
                           sent_today=sent_today,
                           replies_today=replies_today,
                           recent_campaigns=recent_campaigns,
                           recent_replies=recent_replies,
                           daily_stats_labels=daily_stats_labels,
                           daily_stats_values=daily_stats_values,
                           campaign_stats_values=campaign_stats_values)

@app.route('/replies')
@login_required
def replies():
    # Filters
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    server_id = request.args.get('server_id')
    
    query = Reply.query
    
    if start_date:
        query = query.filter(Reply.received_at >= datetime.strptime(start_date, '%Y-%m-%d'))
    if end_date:
        # Add one day to include the end date fully
        end_dt = datetime.strptime(end_date, '%Y-%m-%d')
        query = query.filter(Reply.received_at < end_dt.replace(day=end_dt.day+1))
        
    # If server_id is provided for filtering (optional, if we want to see replies from specific server)
    if server_id:
        query = query.filter_by(server_id=server_id)

    replies_list = query.order_by(Reply.received_at.desc()).all()
    servers = Server.query.all()
    
    return render_template('replies.html', replies=replies_list, servers=servers, 
                           selected_server_id=int(server_id) if server_id else None,
                           start_date=start_date, end_date=end_date)

@app.route('/check_replies_manual', methods=['POST'])
@login_required
def check_replies_manual():
    # Handle both JSON and form data
    if request.is_json:
        data = request.get_json()
        start_date_str = data.get('start_date')
        limit = data.get('limit', 200)
    else:
        start_date_str = request.form.get('start_date')
        limit = request.form.get('limit', 200)

    server = Server.query.filter_by(is_primary=True).first()
    if not server:
        if request.is_json:
            return jsonify({'success': False, 'message': 'No primary server configured.'})
        flash('No primary server configured.', 'error')
        return redirect(url_for('replies'))

    imap_config = {
        'server': server.imap_server,
        'email': server.smtp_email,
        'password': server.smtp_password
    }
    campaigns = Campaign.query.all()
    
    # Parse start date
    start_date = None
    if start_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
        except ValueError:
            pass
            
    # Default to 1st of current month if not provided
    if not start_date:
        today = datetime.now()
        start_date = today.replace(day=1)

    try:
        new_replies, scanned_count, error = check_replies(imap_config, campaigns, start_date, limit)
        if error:
            if request.is_json:
                return jsonify({'success': False, 'message': f"Error checking replies: {error}"})
            flash(f"Error checking replies: {error}", 'error')
        else:
            count = 0
            if new_replies:
                for r in new_replies:
                    exists = Reply.query.filter_by(sender_email=r['sender'], subject=r['subject'], content=r['content']).first()
                    if not exists:
                        reply = Reply(
                            campaign_id=r['campaign_id'],
                            sender_email=r['sender'],
                            subject=r['subject'],
                            content=r['content'],
                            server_id=server.id,
                            cc=r.get('cc'),
                            has_attachments=r.get('has_attachments', False)
                        )
                        db.session.add(reply)
                        count += 1
                db.session.commit()
            
            msg = f'Checked {scanned_count} emails since {start_date.strftime("%Y-%m-%d")}. Found {count} new replies.'
            if request.is_json:
                # Return HTML for table rows to update frontend easily
                # Re-query replies to include new ones and apply filters if needed
                # For simplicity, just returning success and letting frontend reload or we could return data
                # Better: Return the new rows rendered
                return jsonify({'success': True, 'message': msg, 'new_count': count})
            
            flash(msg, 'success')
    except Exception as e:
        if request.is_json:
            return jsonify({'success': False, 'message': f"Error: {str(e)}"})
        flash(f"Error checking replies: {str(e)}", 'error')

    return redirect(url_for('replies'))

@app.route('/replies/<int:reply_id>/resend_campaign', methods=['POST'])
@login_required
def resend_campaign_single(reply_id):
    reply = Reply.query.get_or_404(reply_id)
    campaign = db.session.get(Campaign, reply.campaign_id) if reply.campaign_id else None
    
    if not campaign:
        flash('Associated campaign not found.', 'error')
        return redirect(url_for('replies'))
        
    server = Server.query.filter_by(is_primary=True).first()
    if not server:
        flash('No primary server configured.', 'error')
        return redirect(url_for('replies'))
        
    template = db.session.get(Template, campaign.template_id)
    if not template:
        flash('Campaign template not found.', 'error')
        return redirect(url_for('replies'))
        
    # Try to find contact for personalization
    _, email_addr = parseaddr(reply.sender_email)
    contact = Contact.query.filter_by(email=email_addr).first()
    name = contact.name if contact else 'Valued Customer'
    
    smtp_config = {
        'server': server.smtp_server,
        'port': server.smtp_port,
        'email': server.smtp_email,
        'password': server.smtp_password
    }
    
    personalized_content = template.content.replace('{{name}}', name)
    
    try:
        success, error = send_email(smtp_config, reply.sender_email, template.subject, personalized_content)
        
        log = EmailLog(
            campaign_id=campaign.id,
            recipient_email=reply.sender_email,
            status='sent' if success else 'failed',
            error_message=None if success else error
        )
        db.session.add(log)
        db.session.commit()
        
        if success:
            flash(f'Campaign resent to {reply.sender_email}.', 'success')
        else:
            flash(f'Failed to resend campaign: {error}', 'error')
            
    except Exception as e:
        flash(f'Error resending campaign: {str(e)}', 'error')
        
    return redirect(url_for('replies'))

@app.route('/replies/<int:reply_id>/followup', methods=['POST'])
@login_required
def send_followup(reply_id):
    reply = Reply.query.get_or_404(reply_id)
    subject = request.form.get('subject')
    content = request.form.get('content')
    
    server = Server.query.filter_by(is_primary=True).first()
    if not server:
        flash('No primary server configured.', 'error')
        return redirect(url_for('replies'))
        
    smtp_config = {
        'server': server.smtp_server,
        'port': server.smtp_port,
        'email': server.smtp_email,
        'password': server.smtp_password
    }
    
    try:
        success, error = send_email(smtp_config, reply.sender_email, subject, content)
        
        log = EmailLog(
            campaign_id=reply.campaign_id,
            recipient_email=reply.sender_email,
            status='sent' if success else 'failed',
            type='followup',
            error_message=None if success else error
        )
        db.session.add(log)
        db.session.commit()
        
        if success:
            flash(f'Follow-up sent to {reply.sender_email}.', 'success')
        else:
            flash(f'Failed to send follow-up: {error}', 'error')
            
    except Exception as e:
        flash(f'Error sending follow-up: {str(e)}', 'error')
        
    return redirect(url_for('replies'))

@app.route('/set_primary_server/<int:server_id>')
@login_required
def set_primary_server(server_id):
    # Unset all
    Server.query.update({Server.is_primary: False})
    # Set new primary
    server = Server.query.get_or_404(server_id)
    server.is_primary = True
    db.session.commit()
    flash(f'Primary server set to {server.name}', 'success')
    return redirect(request.referrer or url_for('replies'))

@app.route('/contacts', methods=['GET', 'POST'])
@login_required
def contacts():
    if request.method == 'POST':
        action = request.form.get('action')
        
        if action == 'create_group':
            group_name = request.form.get('group_name')
            if group_name:
                existing = ContactGroup.query.filter_by(name=group_name).first()
                if not existing:
                    group = ContactGroup(name=group_name)
                    db.session.add(group)
                    db.session.commit()
                    flash(f'Group "{group_name}" created.', 'success')
                else:
                    flash('Group already exists.', 'error')
        
        elif action == 'add_to_group':
            group_id = request.form.get('group_id')
            contact_ids = request.form.getlist('contact_ids') # Assuming checkboxes
            if group_id and contact_ids:
                group = db.session.get(ContactGroup, group_id)
                count = 0
                for cid in contact_ids:
                    contact = db.session.get(Contact, cid)
                    if contact and group not in contact.groups:
                        contact.groups.append(group)
                        count += 1
                db.session.commit()
                flash(f'Added {count} contacts to group "{group.name}".', 'success')

        elif action == 'manual_add':
            # Handle manual entry
            manual_entry = request.form.get('manual_entry')
            group_id = request.form.get('group_id') # Optional group assignment on creation
            
            if manual_entry:
                count = 0
                lines = manual_entry.replace(',', '\n').split('\n')
                group = db.session.get(ContactGroup, group_id) if group_id else None
                
                for line in lines:
                    email = line.strip()
                    if email:
                        existing = Contact.query.filter_by(email=email).first()
                        if not existing:
                            contact = Contact(email=email, name='Unknown')
                            if group:
                                contact.groups.append(group)
                            db.session.add(contact)
                            count += 1
                        elif group and group not in existing.groups:
                            existing.groups.append(group)
                            count += 1 # Count as success if added to group
                db.session.commit()
                flash(f'Successfully processed {count} contacts.', 'success')

        elif action == 'import_file':
            # Handle file upload
            file = request.files.get('file')
            group_id = request.form.get('group_id')
            
            if file:
                try:
                    data = []
                    if file.filename.endswith('.xlsx'):
                        wb = load_workbook(file)
                        sheet = wb.active
                        headers = [cell.value for cell in sheet[1]]
                        for row in sheet.iter_rows(min_row=2, values_only=True):
                            data.append(dict(zip(headers, row)))
                    elif file.filename.endswith('.json'):
                        data = json.load(file)
                    else:
                        flash('Invalid file format. Please upload .xlsx or .json', 'error')
                        return redirect(url_for('contacts'))
                    
                    count = 0
                    group = db.session.get(ContactGroup, group_id) if group_id else None
                    
                    for row in data:
                        email = row.get('Email') or row.get('email')
                        if email:
                            existing = Contact.query.filter_by(email=email).first()
                            contact = existing
                            if not existing:
                                contact = Contact(
                                    name=row.get('Name') or row.get('name'),
                                    email=email,
                                    company=row.get('Company') or row.get('company')
                                )
                                db.session.add(contact)
                                count += 1
                            
                            if group and group not in contact.groups:
                                contact.groups.append(group)
                                if existing: count += 1 # Count if just added to group
                                
                    db.session.commit()
                    flash(f'Successfully processed {count} contacts.', 'success')
                except Exception as e:
                    flash(f'Error importing contacts: {str(e)}', 'error')
            
    contacts_list = Contact.query.order_by(Contact.created_at.desc()).all()
    groups = ContactGroup.query.all()
    return render_template('contacts.html', contacts=contacts_list, groups=groups)

@app.route('/templates', methods=['GET', 'POST'])
@login_required
def templates():
    if request.method == 'POST':
        name = request.form.get('name')
        subject = request.form.get('subject')
        content = request.form.get('content')
        
        new_template = Template(name=name, subject=subject, content=content)
        db.session.add(new_template)
        db.session.commit()
        flash('Template created successfully.', 'success')
        return redirect(url_for('templates'))

    templates_list = Template.query.order_by(Template.created_at.desc()).all()
    file_templates = get_file_templates()
    return render_template('templates.html', templates=templates_list, file_templates=file_templates)

@app.route('/templates/file/<filename>')
@login_required
def get_template_content(filename):
    try:
        # Security check: ensure filename doesn't contain path separators
        if '..' in filename or '/' in filename or '\\' in filename:
            return jsonify({'error': 'Invalid filename'}), 400
            
        file_path = os.path.join(TEMPLATE_DIR, filename)
        if not os.path.exists(file_path):
            return jsonify({'error': 'File not found'}), 404
            
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()
            
        return jsonify({'content': content})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/campaigns', methods=['GET', 'POST'])
@login_required
def campaigns():
    if request.method == 'POST':
        name = request.form.get('name')
        template_mode = request.form.get('template_mode')
        target_group_id = request.form.get('target_group_id')
        target_group_id = int(target_group_id) if target_group_id else None
        
        template_id = None
        if template_mode == 'custom':
            custom_subject = request.form.get('custom_subject')
            custom_content = request.form.get('custom_content')
            new_template = Template(name=f"Custom: {name}", subject=custom_subject, content=custom_content)
            db.session.add(new_template)
            db.session.commit()
            template_id = new_template.id
        elif template_mode == 'file':
            file_template = request.form.get('file_template')
            if file_template:
                try:
                    with open(os.path.join(TEMPLATE_DIR, file_template), 'r', encoding='utf-8') as f:
                        content = f.read()
                    new_template = Template(name=f"File: {file_template}", subject=f"Campaign: {name}", content=content)
                    db.session.add(new_template)
                    db.session.commit()
                    template_id = new_template.id
                except Exception as e:
                    flash(f'Error reading template file: {str(e)}', 'error')
                    return redirect(url_for('campaigns'))
        else:
            template_id = request.form.get('template_id')
        
        # Calculate contacts count based on group
        if target_group_id:
            group = db.session.get(ContactGroup, target_group_id)
            contacts_count = len(group.contacts) if group else 0
        else:
            contacts_count = Contact.query.filter_by(status='active').count()
        
        existing = Campaign.query.filter_by(name=name).first()
        if existing:
            flash(f'A campaign with the name "{name}" already exists. Please use a different name.', 'error')
            return redirect(url_for('campaigns'))
        
        new_campaign = Campaign(name=name, template_id=template_id, status='draft', total_contacts=contacts_count, target_group_id=target_group_id)
        db.session.add(new_campaign)
        db.session.commit()
        
        flash(f'Campaign "{name}" created successfully!', 'success')
        return redirect(url_for('campaigns'))

    campaigns_list = Campaign.query.order_by(Campaign.created_at.desc()).all()
    templates_list = Template.query.all()
    file_templates = get_file_templates()
    groups = ContactGroup.query.all()
    return render_template('campaigns.html', campaigns=campaigns_list, templates=templates_list, file_templates=file_templates, groups=groups)

@app.route('/campaigns/<int:campaign_id>/duplicate', methods=['POST'])
@login_required
def duplicate_campaign(campaign_id):
    original = Campaign.query.get_or_404(campaign_id)
    
    # Create new name
    new_name = f"Copy of {original.name}"
    # Ensure unique name
    counter = 1
    while Campaign.query.filter_by(name=new_name).first():
        new_name = f"Copy of {original.name} ({counter})"
        counter += 1
        
    # Recalculate count
    if original.target_group_id:
        group = db.session.get(ContactGroup, original.target_group_id)
        contacts_count = len(group.contacts) if group else 0
    else:
        contacts_count = Contact.query.filter_by(status='active').count()
    
    new_campaign = Campaign(
        name=new_name,
        template_id=original.template_id,
        status='draft',
        total_contacts=contacts_count,
        target_group_id=original.target_group_id
    )
    db.session.add(new_campaign)
    db.session.commit()
    
    flash(f'Campaign duplicated as "{new_name}".', 'success')
    return redirect(url_for('campaigns'))

def send_campaign_emails(campaign_id):
    with app.app_context():
        campaign = db.session.get(Campaign, campaign_id)
        if not campaign:
            return
        
        server = Server.query.filter_by(is_primary=True).first()
        if not server:
            campaign.status = 'failed'
            campaign.error_message = "No primary server configured. Please go to Settings and configure a server."
            db.session.commit()
            print("No primary server configured.")
            return
        
        template = db.session.get(Template, campaign.template_id)
        if not template:
            campaign.status = 'failed'
            campaign.error_message = "Campaign template not found. It may have been deleted."
            db.session.commit()
            return
        
        template_content = template.content
        template_subject = template.subject
        
        # Filter contacts based on target group
        if campaign.target_group_id:
            # Join with groups to filter
            contacts = Contact.query.join(Contact.groups).filter(ContactGroup.id == campaign.target_group_id, Contact.status == 'active').all()
        else:
            contacts = Contact.query.filter_by(status='active').all()
            
        total_contacts = len(contacts)
        campaign.total_contacts = total_contacts
        campaign.status = 'sending'
        campaign.sent_count = 0
        db.session.commit()
        
        smtp_config = {
            'server': server.smtp_server,
            'port': server.smtp_port,
            'email': server.smtp_email,
            'password': server.smtp_password
        }
        
        sent_count = 0
        failed_count = 0
        last_error = None
        
        try:
            for contact in contacts:
                personalized_content = template_content.replace('{{name}}', contact.name or 'Valued Customer')
                
                success, error = send_email(smtp_config, contact.email, template_subject, personalized_content)
                
                log = EmailLog(
                    campaign_id=campaign.id,
                    recipient_email=contact.email,
                    status='sent' if success else 'failed',
                    error_message=None if success else error
                )
                db.session.add(log)
                
                if success:
                    sent_count += 1
                else:
                    failed_count += 1
                    last_error = error
                
                # Commit progress after every email for real-time updates
                campaign.sent_count = sent_count
                db.session.commit()
            
            campaign.sent_count = sent_count
            
            if sent_count == 0 and total_contacts > 0:
                campaign.status = 'failed'
                campaign.error_message = f"All emails failed. Last error: {last_error}"
            else:
                campaign.status = 'completed' if sent_count > 0 or total_contacts == 0 else 'failed'
                
            db.session.commit()
            
        except Exception as e:
            campaign.status = 'failed'
            campaign.error_message = str(e)
            db.session.commit()
            print(f"Error sending campaign {campaign_id}: {str(e)}")

@app.route('/campaigns/<int:campaign_id>/start', methods=['POST'])
@login_required
def start_campaign(campaign_id):
    campaign = Campaign.query.get_or_404(campaign_id)
    
    if campaign.status not in ['draft', 'failed']:
        flash('Campaign can only be started if it is in draft or failed status.', 'error')
        return redirect(url_for('campaigns'))
    
    server = Server.query.filter_by(is_primary=True).first()
    if not server:
        flash('Please configure a primary server in Settings before starting a campaign.', 'error')
        return redirect(url_for('settings'))
    
    if campaign.status == 'failed':
        campaign.sent_count = 0
    
    # Set status to sending immediately so UI updates on redirect
    campaign.status = 'sending'
    db.session.commit()
    
    thread = threading.Thread(target=send_campaign_emails, args=(campaign_id,))
    thread.daemon = True
    thread.start()
    
    flash(f'Campaign "{campaign.name}" started.', 'success')
    return redirect(url_for('campaigns'))

@app.route('/campaigns/<int:campaign_id>/progress', methods=['GET'])
@login_required
def get_campaign_progress(campaign_id):
    campaign = Campaign.query.get_or_404(campaign_id)
    total = campaign.total_contacts if campaign.total_contacts > 0 else 1
    sent = campaign.sent_count
    progress_percent = round((sent / total * 100), 1) if total > 0 else 0
    return jsonify({
        'status': campaign.status,
        'sent': sent,
        'total': total,
        'progress': progress_percent
    })

@app.route('/settings', methods=['GET', 'POST'])
@login_required
def settings():
    if request.method == 'POST':
        # Add new server
        name = request.form.get('name')
        smtp_server = request.form.get('smtp_server')
        smtp_port = int(request.form.get('smtp_port'))
        smtp_email = request.form.get('smtp_email')
        smtp_password = request.form.get('smtp_password')
        imap_server = request.form.get('imap_server')
        
        # If this is the first server, make it primary
        is_primary = Server.query.count() == 0
        
        new_server = Server(
            name=name,
            smtp_server=smtp_server,
            smtp_port=smtp_port,
            smtp_email=smtp_email,
            smtp_password=smtp_password,
            imap_server=imap_server,
            is_primary=is_primary
        )
        db.session.add(new_server)
        db.session.commit()
        flash('Server added successfully.', 'success')
        return redirect(url_for('settings'))
        
    servers = Server.query.all()
    return render_template('settings.html', servers=servers)

@app.route('/settings/upload_db', methods=['POST'])
@login_required
def upload_db():
    if 'file' not in request.files:
        flash('No file part', 'error')
        return redirect(url_for('settings'))
    file = request.files['file']
    if file.filename == '':
        flash('No selected file', 'error')
        return redirect(url_for('settings'))
    
    if file and (file.filename.endswith('.db') or file.filename.endswith('.sqlite')):
        try:
            # Close DB session to release locks
            db.session.remove()
            
            # Determine path
            # Assuming standard Flask instance path or relative to app
            instance_path = os.path.join(basedir, 'instance')
            if not os.path.exists(instance_path):
                os.makedirs(instance_path)
                
            db_path = os.path.join(instance_path, 'email_marketing.db')
            backup_path = db_path + '.bak'
            
            # Backup existing
            if os.path.exists(db_path):
                shutil.copy2(db_path, backup_path)
            
            # Save new file
            file.save(db_path)
            
            flash('Database restored successfully. The application state has been updated.', 'success')
        except Exception as e:
            flash(f'Error restoring database: {str(e)}', 'error')
    else:
        flash('Invalid file type. Please upload a .db or .sqlite file.', 'error')
        
    return redirect(url_for('settings'))

@app.route('/settings/change_password', methods=['POST'])
@login_required
def change_password():
    old_password = request.form.get('old_password')
    new_password = request.form.get('new_password')
    confirm_password = request.form.get('confirm_password')

    if not old_password or not new_password or not confirm_password:
        return jsonify({'success': False, 'message': 'All fields are required.'})

    if current_user.password != old_password:
        return jsonify({'success': False, 'message': 'Incorrect old password.'})

    if new_password != confirm_password:
        return jsonify({'success': False, 'message': 'New passwords do not match.'})

    current_user.password = new_password
    db.session.commit()
    return jsonify({'success': True, 'message': 'Password updated successfully.'})

@app.route('/contacts/<int:contact_id>/delete', methods=['POST'])
@login_required
def delete_contact(contact_id):
    contact = Contact.query.get_or_404(contact_id)
    db.session.delete(contact)
    db.session.commit()
    flash('Contact deleted successfully.', 'success')
    return redirect(url_for('contacts'))

@app.route('/contacts/<int:contact_id>/edit', methods=['POST'])
@login_required
def edit_contact(contact_id):
    contact = Contact.query.get_or_404(contact_id)
    contact.name = request.form.get('name')
    contact.email = request.form.get('email')
    contact.company = request.form.get('company')
    db.session.commit()
    flash('Contact updated successfully.', 'success')
    return redirect(url_for('contacts'))

@app.route('/templates/<int:template_id>/delete', methods=['POST'])
@login_required
def delete_template(template_id):
    template = Template.query.get_or_404(template_id)
    # Check if used in any campaign
    if Campaign.query.filter_by(template_id=template_id).first():
        flash('Cannot delete template because it is used in a campaign.', 'error')
    else:
        db.session.delete(template)
        db.session.commit()
        flash('Template deleted successfully.', 'success')
    return redirect(url_for('templates'))

@app.route('/templates/<int:template_id>/edit', methods=['POST'])
@login_required
def edit_template(template_id):
    template = Template.query.get_or_404(template_id)
    template.name = request.form.get('name')
    template.subject = request.form.get('subject')
    template.content = request.form.get('content')
    db.session.commit()
    flash('Template updated successfully.', 'success')
    return redirect(url_for('templates'))

@app.route('/campaigns/<int:campaign_id>/delete', methods=['POST'])
@login_required
def delete_campaign(campaign_id):
    campaign = Campaign.query.get_or_404(campaign_id)
    # Delete associated logs and replies first if needed, or rely on cascade if configured (not configured here)
    # Manually delete logs and replies to be safe
    EmailLog.query.filter_by(campaign_id=campaign_id).delete()
    Reply.query.filter_by(campaign_id=campaign_id).delete()
    db.session.delete(campaign)
    db.session.commit()
    flash('Campaign deleted successfully.', 'success')
    return redirect(url_for('campaigns'))

@app.route('/campaigns/<int:campaign_id>/edit', methods=['POST'])
@login_required
def edit_campaign(campaign_id):
    campaign = Campaign.query.get_or_404(campaign_id)
    if campaign.status != 'draft':
        flash('Only draft campaigns can be edited.', 'error')
        return redirect(url_for('campaigns'))
    
    campaign.name = request.form.get('name')
    
    # Update Target Group
    target_group_id = request.form.get('target_group_id')
    campaign.target_group_id = int(target_group_id) if target_group_id else None
    
    # Update Template
    template_id = request.form.get('template_id')
    if template_id:
        campaign.template_id = int(template_id)
        
    # Recalculate total contacts
    if campaign.target_group_id:
        group = db.session.get(ContactGroup, campaign.target_group_id)
        campaign.total_contacts = len(group.contacts) if group else 0
    else:
        campaign.total_contacts = Contact.query.filter_by(status='active').count()
        
    db.session.commit()
    flash('Campaign updated successfully.', 'success')
    return redirect(url_for('campaigns'))

@app.route('/settings/server/<int:server_id>/delete', methods=['POST'])
@login_required
def delete_server(server_id):
    server = Server.query.get_or_404(server_id)
    if server.is_primary:
        flash('Cannot delete the primary server. Please set another server as primary first.', 'error')
    else:
        # Check if used in replies
        if Reply.query.filter_by(server_id=server_id).first():
             flash('Cannot delete server because it has associated replies. Please clear replies first.', 'error')
        else:
            db.session.delete(server)
            db.session.commit()
            flash('Server deleted successfully.', 'success')
    return redirect(url_for('settings'))

@app.route('/settings/server/<int:server_id>/edit', methods=['POST'])
@login_required
def edit_server(server_id):
    server = Server.query.get_or_404(server_id)
    server.name = request.form.get('name')
    server.smtp_server = request.form.get('smtp_server')
    server.smtp_port = int(request.form.get('smtp_port'))
    server.smtp_email = request.form.get('smtp_email')
    if request.form.get('smtp_password'):
        server.smtp_password = request.form.get('smtp_password')
    server.imap_server = request.form.get('imap_server')
    
    db.session.commit()
    flash('Server updated successfully.', 'success')
    return redirect(url_for('settings'))

@app.route('/settings/download_db')
@login_required
def download_db():
    try:
        instance_path = os.path.join(basedir, 'instance')
        db_path = os.path.join(instance_path, 'email_marketing.db')
        
        if not os.path.exists(db_path):
            # Fallback to root if not in instance
            db_path = os.path.join(basedir, 'email_marketing.db')
            
        if os.path.exists(db_path):
            return send_file(db_path, as_attachment=True, download_name='email_marketing_backup.db')
        else:
            flash('Database file not found.', 'error')
            return redirect(url_for('settings'))
    except Exception as e:
        flash(f'Error downloading database: {str(e)}', 'error')
        return redirect(url_for('settings'))

@app.route('/settings/server/<int:server_id>/check_status')
@login_required
def check_server_status(server_id):
    server = Server.query.get_or_404(server_id)
    
    smtp_status = False
    imap_status = False
    error_message = ""
    
    # Check SMTP
    try:
        import smtplib
        if server.smtp_port == 465:
            smtp = smtplib.SMTP_SSL(server.smtp_server, server.smtp_port, timeout=15)
        else:
            smtp = smtplib.SMTP(server.smtp_server, server.smtp_port, timeout=15)
            smtp.starttls()
            
        smtp.login(server.smtp_email, server.smtp_password)
        smtp.quit()
        smtp_status = True
    except Exception as e:
        error_message += f"SMTP Error: {str(e)}. "

    # Check IMAP
    try:
        import imaplib
        # Some servers don't support IMAP, so if the server is clearly not an IMAP server (like smtp-relay), this will fail.
        # We'll try to connect.
        if server.imap_server and server.imap_server.strip():
            imap = imaplib.IMAP4_SSL(server.imap_server, timeout=15)
            imap.login(server.smtp_email, server.smtp_password)
            imap.logout()
            imap_status = True
        else:
            # If no IMAP server provided (or empty), consider it skipped/OK or just False but not an error
            pass
    except Exception as e:
        error_message += f"IMAP Error: {str(e)}."
        
    # Determine overall status
    # If SMTP works, we consider it a success for sending purposes.
    # IMAP failure is a warning.
    
    success = smtp_status
    if smtp_status and not imap_status:
        error_message += " (Warning: SMTP connected but IMAP failed. You can send emails but replies won't be tracked.)"
    
    return jsonify({
        'success': success, 
        'smtp_status': smtp_status,
        'imap_status': imap_status,
        'message': error_message.strip()
    })


if __name__ == '__main__':
    app.run(debug=False)
